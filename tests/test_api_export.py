from __future__ import annotations

import csv
import io

import openpyxl
import pytest

BASE_AMOUNT = 6_543_210.0


def _post_score(client, **overrides):
    payload = {
        "step": 3, "type": "TRANSFER", "amount": BASE_AMOUNT,
        "oldbalanceOrg": BASE_AMOUNT, "newbalanceOrig": 0.0,
        "oldbalanceDest": 0.0, "newbalanceDest": 0.0,
    }
    payload.update(overrides)
    res = client.post("/score", json=payload)
    assert res.status_code == 200
    return res.json()


@pytest.fixture(scope="module")
def seeded(client):
    t1 = _post_score(client, amount=BASE_AMOUNT + 1, oldbalanceOrg=BASE_AMOUNT + 1, login_country="TR")
    t2 = _post_score(client, amount=BASE_AMOUNT + 2, oldbalanceOrg=BASE_AMOUNT + 2, login_country="TR")
    t3 = _post_score(client, amount=BASE_AMOUNT + 3, oldbalanceOrg=BASE_AMOUNT + 3, login_country="TR")
    assert t1["risk_band"] == t2["risk_band"] == t3["risk_band"] == "RED"
    return {"t1": t1, "t2": t2, "t3": t3}


def _filter_params():
    return {"risk_band": "RED", "amount_min": BASE_AMOUNT, "amount_max": BASE_AMOUNT + 3}


def test_export_requires_csv_or_xlsx_format(client, seeded):
    res = client.get("/cases/export", params={**_filter_params(), "format": "pdf"})
    assert res.status_code == 400


def test_csv_export_matches_the_equivalent_list_endpoint(client, seeded):
    list_res = client.get("/cases", params=_filter_params())
    csv_res = client.get("/cases/export", params={**_filter_params(), "format": "csv"})
    assert csv_res.status_code == 200
    assert csv_res.headers["content-type"].startswith("text/csv")
    assert "attachment" in csv_res.headers["content-disposition"]

    rows = list(csv.DictReader(io.StringIO(csv_res.text)))
    assert len(rows) == list_res.json()["total"]
    csv_txn_ids = {int(r["transaction_id"]) for r in rows}
    list_txn_ids = {item["transaction_id"] for item in list_res.json()["items"]}
    assert csv_txn_ids == list_txn_ids


def test_xlsx_export_row_count_matches_csv_and_list(client, seeded):
    list_res = client.get("/cases", params=_filter_params())
    xlsx_res = client.get("/cases/export", params={**_filter_params(), "format": "xlsx"})
    assert xlsx_res.status_code == 200
    assert xlsx_res.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_res.content))
    ws = wb["Cases"]
    data_row_count = ws.max_row - 2
    assert data_row_count == list_res.json()["total"]

    xlsx_txn_ids = {ws.cell(row=r, column=2).value for r in range(3, ws.max_row + 1)}
    list_txn_ids = {item["transaction_id"] for item in list_res.json()["items"]}
    assert xlsx_txn_ids == list_txn_ids


def test_csv_and_xlsx_agree_on_exact_row_set(client, seeded):
    csv_res = client.get("/cases/export", params={**_filter_params(), "format": "csv"})
    xlsx_res = client.get("/cases/export", params={**_filter_params(), "format": "xlsx"})

    csv_ids = {int(r["transaction_id"]) for r in csv.DictReader(io.StringIO(csv_res.text))}
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_res.content))
    ws = wb["Cases"]
    xlsx_ids = {ws.cell(row=r, column=2).value for r in range(3, ws.max_row + 1)}

    assert csv_ids == xlsx_ids


def test_csv_source_column_marks_api_posted_rows_as_live(client, seeded):
    csv_res = client.get("/cases/export", params={**_filter_params(), "format": "csv"})
    rows = list(csv.DictReader(io.StringIO(csv_res.text)))
    assert all(r["source"] == "live" for r in rows)


def test_export_respects_filters_not_just_total_dataset(client, seeded):
    res = client.get("/cases/export", params={**_filter_params(), "country": "DE", "format": "csv"})
    rows = list(csv.DictReader(io.StringIO(res.text)))
    assert len(rows) == 0


def test_csv_columns_match_expected_schema(client, seeded):
    res = client.get("/cases/export", params={**_filter_params(), "format": "csv"})
    reader = csv.DictReader(io.StringIO(res.text))
    assert reader.fieldnames == [
        "case_id", "transaction_id", "status", "risk_band", "hybrid_score",
        "type", "amount", "login_country", "source", "decision", "created_at",
    ]


def _empty_filter_params():
    return {**_filter_params(), "country": "ZZ"}


def test_csv_export_on_empty_result_is_a_valid_header_only_file(client, seeded):
    res = client.get("/cases/export", params={**_empty_filter_params(), "format": "csv"})
    assert res.status_code == 200
    rows = list(csv.DictReader(io.StringIO(res.text)))
    assert rows == []
    assert res.text.strip().split("\r\n")[0].split(",")[0] == "case_id" or res.text.strip().split("\n")[0].split(",")[0] == "case_id"


def test_xlsx_export_on_empty_result_opens_and_has_correct_structure(client, seeded):
    res = client.get("/cases/export", params={**_empty_filter_params(), "format": "xlsx"})
    assert res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb["Cases"]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "case_id"

    ws2 = wb["Summary"]
    assert ws2["B5"].value == 0
    assert len(ws2._charts) == 0


def test_xlsx_export_with_exactly_one_row_still_builds_a_chart(client, seeded):
    res = client.get(
        "/cases/export",
        params={"amount_min": BASE_AMOUNT + 1, "amount_max": BASE_AMOUNT + 1, "format": "xlsx"},
    )
    assert res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb["Cases"]
    assert ws.max_row == 3
    ws2 = wb["Summary"]
    assert ws2["B5"].value == 1
    assert len(ws2._charts) == 1
